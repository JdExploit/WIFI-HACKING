#!/usr/bin/env python3
"""
Script: wifi_security_analyzer.py
Descripción: Analizador automatizado de seguridad Wi-Fi WPA2/WPA3 con salida CSV
Autor: Asistente WiFi Security Lab
Versión: 1.1 (con CSV output)
"""

import subprocess
import json
import csv
import re
from datetime import datetime
import os
import sys
from typing import List, Dict, Tuple, Optional
import pandas as pd

class WiFiSecurityAnalyzer:
    def __init__(self, interface: str = "wlan0mon"):
        """
        Inicializa el analizador de seguridad Wi-Fi
        
        Args:
            interface: Interfaz en modo monitor
        """
        self.interface = interface
        self.results = {
            "timestamp": datetime.now().isoformat(),
            "interface": interface,
            "networks": [],
            "analysis": {}
        }
        
    def check_root(self) -> bool:
        """Verifica si se ejecuta como root"""
        return os.geteuid() == 0
    
    def run_command(self, cmd: str) -> Tuple[str, str]:
        """
        Ejecuta un comando y retorna stdout y stderr
        
        Args:
            cmd: Comando a ejecutar
            
        Returns:
            Tuple (stdout, stderr)
        """
        try:
            result = subprocess.run(
                cmd,
                shell=True,
                capture_output=True,
                text=True,
                timeout=30
            )
            return result.stdout, result.stderr
        except subprocess.TimeoutExpired:
            return "", "Timeout expired"
        except Exception as e:
            return "", str(e)
    
    def scan_networks(self, duration: int = 60) -> List[Dict]:
        """
        Escanea redes Wi-Fi disponibles y retorna en formato CSV parseable
        
        Args:
            duration: Duración del escaneo en segundos
            
        Returns:
            Lista de redes detectadas
        """
        print(f"[*] Escaneando redes durante {duration} segundos...")
        
        # Capturar redes con airodump-ng (genera CSV automáticamente)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f"scan_{timestamp}"
        cmd = f"timeout {duration} sudo airodump-ng {self.interface} --output-format csv -w {output_file}"
        stdout, stderr = self.run_command(cmd)
        
        networks = []
        
        # Leer archivo CSV generado por airodump-ng
        csv_file = f"{output_file}-01.csv"
        if os.path.exists(csv_file):
            print(f"[+] Archivo CSV generado: {csv_file}")
            
            # Parsear CSV de airodump-ng
            with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
                
                # Separar secciones (redes y clientes)
                sections = content.split('\n\n')
                
                if len(sections) >= 1:
                    # Parsear sección de redes
                    network_lines = sections[0].strip().split('\n')
                    
                    # Encontrar encabezados
                    for i, line in enumerate(network_lines):
                        if "BSSID" in line and "ESSID" in line:
                            headers = [h.strip() for h in line.split(',')]
                            start_idx = i + 1
                            break
                    
                    # Parsear datos de redes
                    for line in network_lines[start_idx:]:
                        if line.strip() and "Station MAC" not in line:
                            values = [v.strip() for v in line.split(',')]
                            if len(values) >= 14:
                                network = {
                                    "bssid": values[0] if len(values) > 0 else "",
                                    "first_time_seen": values[1] if len(values) > 1 else "",
                                    "last_time_seen": values[2] if len(values) > 2 else "",
                                    "channel": values[3] if len(values) > 3 else "",
                                    "speed": values[4] if len(values) > 4 else "",
                                    "privacy": values[5] if len(values) > 5 else "",
                                    "cipher": values[6] if len(values) > 6 else "",
                                    "authentication": values[7] if len(values) > 7 else "",
                                    "power": values[8] if len(values) > 8 else "",
                                    "beacons": values[9] if len(values) > 9 else "",
                                    "iv": values[10] if len(values) > 10 else "",
                                    "lan_ip": values[11] if len(values) > 11 else "",
                                    "id_length": values[12] if len(values) > 12 else "",
                                    "essid": values[13] if len(values) > 13 else "(hidden)",
                                    "key": values[14] if len(values) > 14 else ""
                                }
                                networks.append(network)
            
            print(f"[+] {len(networks)} redes parseadas del CSV")
            
            # Guardar CSV limpio para el usuario
            self._save_clean_csv(networks, f"networks_scan_{timestamp}.csv")
        
        self.results["networks"] = networks
        return networks
    
    def _save_clean_csv(self, networks: List[Dict], filename: str):
        """
        Guarda los datos de redes en un CSV limpio y estructurado
        
        Args:
            networks: Lista de redes
            filename: Nombre del archivo CSV
        """
        if not networks:
            return
            
        # Definir columnas para el CSV
        fieldnames = [
            "ESSID", "BSSID", "Channel", "Privacy", "Cipher", 
            "Authentication", "Power", "Beacons", "IV", "First Seen",
            "Last Seen", "Speed", "LAN IP"
        ]
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            for net in networks:
                writer.writerow({
                    "ESSID": net.get("essid", ""),
                    "BSSID": net.get("bssid", ""),
                    "Channel": net.get("channel", ""),
                    "Privacy": net.get("privacy", ""),
                    "Cipher": net.get("cipher", ""),
                    "Authentication": net.get("authentication", ""),
                    "Power": net.get("power", ""),
                    "Beacons": net.get("beacons", ""),
                    "IV": net.get("iv", ""),
                    "First Seen": net.get("first_time_seen", ""),
                    "Last Seen": net.get("last_time_seen", ""),
                    "Speed": net.get("speed", ""),
                    "LAN IP": net.get("lan_ip", "")
                })
        
        print(f"[+] CSV limpio guardado: {filename}")
    
    def analyze_wpa2_security(self, network: Dict) -> Dict:
        """
        Analiza seguridad WPA2 de una red
        
        Args:
            network: Información de la red
            
        Returns:
            Dict con análisis de seguridad
        """
        analysis = {
            "wpa_version": None,
            "pmf_status": None,
            "cipher_strength": None,
            "security_score": 0,
            "vulnerabilities": [],
            "recommendations": []
        }
        
        privacy = network.get("privacy", "").upper()
        cipher = network.get("cipher", "").upper()
        auth = network.get("authentication", "").upper()
        
        # Determinar versión WPA
        if "WPA3" in privacy:
            analysis["wpa_version"] = "WPA3"
            analysis["security_score"] += 3
        elif "WPA2" in privacy:
            analysis["wpa_version"] = "WPA2"
            analysis["security_score"] += 2
        elif "WPA" in privacy:
            analysis["wpa_version"] = "WPA"
            analysis["security_score"] += 1
        elif "WEP" in privacy:
            analysis["wpa_version"] = "WEP"
            analysis["security_score"] -= 2
        elif "OPN" in privacy:
            analysis["wpa_version"] = "OPEN"
            analysis["security_score"] -= 3
        
        # Analizar cifrado
        if "CCMP" in cipher:
            analysis["cipher_strength"] = "Strong"
            analysis["security_score"] += 2
        elif "TKIP" in cipher:
            analysis["cipher_strength"] = "Weak"
            analysis["security_score"] -= 1
            analysis["vulnerabilities"].append("TKIP vulnerable")
            analysis["recommendations"].append("Upgrade to AES-CCMP")
        
        # Analizar autenticación
        if "SAE" in auth:
            analysis["pmf_status"] = "Required"
            analysis["security_score"] += 2
        elif "PSK" in auth:
            analysis["pmf_status"] = self.check_pmf_support(network["bssid"])
            if analysis["pmf_status"] == "Required":
                analysis["security_score"] += 2
            elif analysis["pmf_status"] == "Optional":
                analysis["security_score"] += 1
        
        # Clasificación de seguridad
        if analysis["security_score"] >= 4:
            analysis["security_level"] = "High"
        elif analysis["security_score"] >= 2:
            analysis["security_level"] = "Medium"
        elif analysis["security_score"] >= 0:
            analysis["security_level"] = "Low"
        else:
            analysis["security_level"] = "Critical"
        
        return analysis
    
    def check_pmf_support(self, bssid: str) -> str:
        """
        Verifica soporte de Management Frame Protection (PMF)
        
        Args:
            bssid: Dirección MAC del AP
            
        Returns:
            Estado de PMF
        """
        # Método simplificado para ejemplo
        # En implementación real, analizaría beacon frames
        return "Unknown"
    
    def generate_csv_report(self, filename: str = None):
        """
        Genera reporte CSV completo del análisis
        
        Args:
            filename: Nombre del archivo CSV (opcional)
        """
        if not self.results["networks"]:
            print("[-] No hay datos para generar CSV")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wifi_analysis_{timestamp}.csv"
        
        # Preparar datos para CSV
        csv_data = []
        for network in self.results["networks"]:
            analysis = self.analyze_wpa2_security(network)
            
            row = {
                # Datos básicos
                "ESSID": network.get("essid", ""),
                "BSSID": network.get("bssid", ""),
                "Channel": network.get("channel", ""),
                "Power_dBm": network.get("power", ""),
                
                # Seguridad
                "Privacy": network.get("privacy", ""),
                "Cipher": network.get("cipher", ""),
                "Authentication": network.get("authentication", ""),
                
                # Análisis
                "WPA_Version": analysis.get("wpa_version", ""),
                "PMF_Status": analysis.get("pmf_status", ""),
                "Cipher_Strength": analysis.get("cipher_strength", ""),
                "Security_Score": analysis.get("security_score", 0),
                "Security_Level": analysis.get("security_level", ""),
                
                # Estadísticas
                "Beacons": network.get("beacons", ""),
                "IV_Count": network.get("iv", ""),
                "First_Seen": network.get("first_time_seen", ""),
                "Last_Seen": network.get("last_time_seen", ""),
                
                # Vulnerabilidades (como string separado por |)
                "Vulnerabilities": "|".join(analysis.get("vulnerabilities", [])),
                "Recommendations": "|".join(analysis.get("recommendations", []))
            }
            csv_data.append(row)
        
        # Escribir CSV
        fieldnames = [
            "ESSID", "BSSID", "Channel", "Power_dBm",
            "Privacy", "Cipher", "Authentication",
            "WPA_Version", "PMF_Status", "Cipher_Strength",
            "Security_Score", "Security_Level",
            "Beacons", "IV_Count", "First_Seen", "Last_Seen",
            "Vulnerabilities", "Recommendations"
        ]
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(csv_data)
        
        print(f"[+] Reporte CSV generado: {filename}")
        print(f"    Total registros: {len(csv_data)}")
        
        # También generar resumen estadístico
        self._generate_statistics_csv(csv_data, filename.replace('.csv', '_stats.csv'))
    
    def _generate_statistics_csv(self, data: List[Dict], filename: str):
        """
        Genera CSV con estadísticas del análisis
        
        Args:
            data: Datos del análisis
            filename: Nombre del archivo
        """
        if not data:
            return
        
        stats = {
            "total_networks": len(data),
            "wpa3_count": sum(1 for d in data if d.get("WPA_Version") == "WPA3"),
            "wpa2_count": sum(1 for d in data if d.get("WPA_Version") == "WPA2"),
            "open_count": sum(1 for d in data if d.get("WPA_Version") == "OPEN"),
            "wep_count": sum(1 for d in data if d.get("WPA_Version") == "WEP"),
            "high_security": sum(1 for d in data if d.get("Security_Level") == "High"),
            "medium_security": sum(1 for d in data if d.get("Security_Level") == "Medium"),
            "low_security": sum(1 for d in data if d.get("Security_Level") == "Low"),
            "critical_security": sum(1 for d in data if d.get("Security_Level") == "Critical"),
            "avg_security_score": sum(d.get("Security_Score", 0) for d in data) / len(data) if data else 0
        }
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Metric", "Value"])
            for key, value in stats.items():
                writer.writerow([key.replace('_', ' ').title(), value])
        
        print(f"[+] Estadísticas CSV generado: {filename}")
    
    def generate_excel_report(self, filename: str = None):
        """
        Genera reporte en Excel (requiere pandas)
        
        Args:
            filename: Nombre del archivo Excel
        """
        try:
            import pandas as pd
        except ImportError:
            print("[-] Pandas no instalado. Instala con: pip install pandas")
            return
        
        if not self.results["networks"]:
            print("[-] No hay datos para generar Excel")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wifi_analysis_{timestamp}.xlsx"
        
        # Preparar datos
        data_for_excel = []
        for network in self.results["networks"]:
            analysis = self.analyze_wpa2_security(network)
            
            row = {
                "ESSID": network.get("essid", ""),
                "BSSID": network.get("bssid", ""),
                "Channel": network.get("channel", ""),
                "Signal (dBm)": network.get("power", ""),
                "Encryption": network.get("privacy", ""),
                "Cipher": network.get("cipher", ""),
                "Auth": network.get("authentication", ""),
                "WPA Version": analysis.get("wpa_version", ""),
                "PMF": analysis.get("pmf_status", ""),
                "Security Level": analysis.get("security_level", ""),
                "Security Score": analysis.get("security_score", 0),
                "Vulnerabilities": "\n".join(analysis.get("vulnerabilities", [])),
                "Recommendations": "\n".join(analysis.get("recommendations", []))
            }
            data_for_excel.append(row)
        
        # Crear DataFrame y guardar Excel
        df = pd.DataFrame(data_for_excel)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Networks', index=False)
            
            # Agregar hoja de resumen
            summary_data = {
                'Metric': ['Total Networks', 'WPA3 Networks', 'WPA2 Networks', 
                          'Open Networks', 'Average Security Score'],
                'Value': [
                    len(df),
                    len(df[df['WPA Version'] == 'WPA3']),
                    len(df[df['WPA Version'] == 'WPA2']),
                    len(df[df['Encryption'].str.contains('OPN', na=False)]),
                    df['Security Score'].mean()
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Formato condicional (ejemplo básico)
            workbook = writer.book
            worksheet = writer.sheets['Networks']
            
            # Resaltar por nivel de seguridad
            from openpyxl.styles import PatternFill
            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            
            for row in range(2, len(df) + 2):  # +2 por encabezado y base 1
                cell = worksheet[f'J{row}']  # Columna J = Security Level
                if cell.value == "High":
                    cell.fill = green_fill
                elif cell.value == "Medium":
                    cell.fill = yellow_fill
                elif cell.value in ["Low", "Critical"]:
                    cell.fill = red_fill
        
        print(f"[+] Reporte Excel generado: {filename}")
    
    def run_complete_analysis(self, output_formats: List[str] = None):
        """
        Ejecuta análisis completo con múltiples formatos de salida
        
        Args:
            output_formats: Lista de formatos (csv, excel, json, html, txt)
        """
        if output_formats is None:
            output_formats = ["csv", "txt"]
        
        print("=" * 60)
        print("     ANALIZADOR WI-FI CON SALIDA CSV/EXCEL")
        print("=" * 60)
        
        # Verificar permisos
        if not self.check_root():
            print("[-] ERROR: Se requieren permisos de root")
            print("    Ejecuta: sudo python3 wifi_security_analyzer.py")
            sys.exit(1)
        
        # Escanear redes
        networks = self.scan_networks(duration=30)
        
        if not networks:
            print("[-] No se detectaron redes Wi-Fi")
            return
        
        # Analizar cada red
        print(f"\n[*] Analizando {len(networks)} redes...")
        for network in networks:
            network["security_analysis"] = self.analyze_wpa2_security(network)
        
        # Generar reportes en formatos solicitados
        print("\n[*] Generando reportes...")
        
        if "csv" in output_formats:
            self.generate_csv_report()
        
        if "excel" in output_formats:
            self.generate_excel_report()
        
        if "json" in output_formats:
            self._generate_json_report()
        
        if "html" in output_formats:
            self._generate_html_report()
        
        if "txt" in output_formats:
            self._generate_text_report()
        
        print("\n" + "=" * 60)
        print("[+] Análisis completado")
        print("=" * 60)
        
        # Mostrar vista previa
        self._show_preview(networks[:5])  # Primeras 5 redes
    
    def _generate_json_report(self, filename: str = None):
        """Genera reporte JSON"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wifi_analysis_{timestamp}.json"
        
        with open(filename, 'w') as f:
            json.dump(self.results, f, indent=2)
        print(f"[+] Reporte JSON: {filename}")
    
    def _generate_html_report(self, filename: str = None):
        """Genera reporte HTML (simplificado)"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wifi_analysis_{timestamp}.html"
        
        html = f"""
        <html>
        <head><title>Wi-Fi Analysis Report</title></head>
        <body>
            <h1>Wi-Fi Security Analysis Report</h1>
            <p>Generated: {datetime.now()}</p>
            <p>Total networks: {len(self.results['networks'])}</p>
        </body>
        </html>
        """
        
        with open(filename, 'w') as f:
            f.write(html)
        print(f"[+] Reporte HTML: {filename}")
    
    def _generate_text_report(self, filename: str = None):
        """Genera reporte de texto"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"wifi_analysis_{timestamp}.txt"
        
        with open(filename, 'w') as f:
            f.write(f"Wi-Fi Analysis Report\n")
            f.write(f"Generated: {datetime.now()}\n")
            f.write(f"Total networks: {len(self.results['networks'])}\n\n")
            
            for net in self.results['networks']:
                f.write(f"ESSID: {net.get('essid')}\n")
                f.write(f"BSSID: {net.get('bssid')}\n")
                f.write(f"Security: {net.get('privacy')}\n\n")
        
        print(f"[+] Reporte de texto: {filename}")
    
    def _show_preview(self, networks: List[Dict]):
        """Muestra vista previa de los resultados"""
        print("\n📊 VISTA PREVIA (primeras 5 redes):")
        print("-" * 80)
        print(f"{'ESSID':<20} {'BSSID':<20} {'Channel':<8} {'Security':<12} {'Score':<6}")
        print("-" * 80)
        
        for net in networks:
            essid = net.get('essid', '')[:18] + '..' if len(net.get('essid', '')) > 18 else net.get('essid', '')
            analysis = net.get('security_analysis', {})
            print(f"{essid:<20} {net.get('bssid', '')[:18]:<20} "
                  f"{net.get('channel', ''):<8} "
                  f"{analysis.get('wpa_version', 'Unknown'):<12} "
                  f"{analysis.get('security_score', 0):<6}")


# Función principal con argumentos
def main():
    """Función principal"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Wi-Fi Security Analyzer')
    parser.add_argument('--interface', '-i', default='wlan0mon', help='Interfaz en modo monitor')
    parser.add_argument('--duration', '-d', type=int, default=30, help='Duración del escaneo (segundos)')
    parser.add_argument('--format', '-f', action='append', 
                       choices=['csv', 'excel', 'json', 'html', 'txt', 'all'],
                       help='Formatos de salida')
    parser.add_argument('--output', '-o', help='Nombre base de archivos de salida')
    
    args = parser.parse_args()
    
    # Determinar formatos
    if args.format:
        if 'all' in args.format:
            formats = ['csv', 'excel', 'json', 'html', 'txt']
        else:
            formats = args.format
    else:
        formats = ['csv', 'txt']
    
    # Crear y ejecutar analizador
    analyzer = WiFiSecurityAnalyzer(interface=args.interface)
    analyzer.run_complete_analysis(output_formats=formats)


if __name__ == "__main__":
    main()
